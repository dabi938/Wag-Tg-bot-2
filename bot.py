from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
import os
import openpyxl

# Define the bot owner's Telegram user ID (replace with your actual user ID)
OWNER_USER_ID = 6742597078  # Replace with the owner's user ID

# Handle file upload
async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Check if the user is the owner
    if update.message.from_user.id != OWNER_USER_ID:
        await update.message.reply_text("Sorry, only the owner can upload files.")
        return

    file = update.message.document
    if file.file_name.endswith('.xlsx'):
        file_path = f"downloads/{file.file_name}"
        os.makedirs('downloads', exist_ok=True)  # Ensure directory exists
        try:
            file_id = file.file_id
            new_file = await context.bot.get_file(file_id)
            await new_file.download_to_drive(file_path)

            await update.message.reply_text(
                f"File {file.file_name} received and saved. The bot will now send grades to the users."
            )
            context.user_data['file_path'] = file_path
            await send_grades_to_users(file_path, update, context)
        except Exception as e:
            await update.message.reply_text(f"Error saving file: {str(e)}")
    else:
        await update.message.reply_text("Sorry, only Excel file is allowed.")

# Function to send grades to users based on their user ID in the Excel file
async def send_grades_to_users(file_path, update, context):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Iterate over the rows to get the users' Telegram user IDs and grades
    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header row
        name = sheet.cell(row=row, column=1).value
        user_id = sheet.cell(row=row, column=2).value  # Assume column 2 contains user IDs
        user_class = sheet.cell(row=row, column=3).value
        grade = sheet.cell(row=row, column=4).value

        # Log user_id for debugging
        print(f"Attempting to send grade to user_id: {user_id} for {name}")

        # Send the grade to the user
        try:
            await context.bot.send_message(
                chat_id=user_id,  # Send message using user ID
                text=f"Hello {name}, your grade in {user_class} is: {grade}"
            )
        except Exception as e:
            # Handle and log errors more clearly
            error_message = f"Error sending grade to {name}: {str(e)}"
            await update.message.reply_text(error_message)
            print(f"Error for {name}: {str(e)}")  # Log error for further inspection

    workbook.close()

# Handle user messages for invalid input
async def handle_invalid_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.from_user.id == OWNER_USER_ID:
        # Owner should only send Excel files
        await update.message.reply_text("Sorry, only Excel file is allowed.")
    else:
        # Users should not be sending anything
        await update.message.reply_text("Sorry, this bot is only for receiving, not for sending.")

# Cancel the conversation
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Operation cancelled.")

# Main function to start the bot
def main():
    TOKEN = "7761836422:AAGCkUlpzvpB5Htstl7y9Jxt0AA7drOjnac"  # Replace with your bot token

    # Create the Application using the default settings
    app = Application.builder().token(TOKEN).build()

    # Add handlers for invalid input
    app.add_handler(MessageHandler(filters.TEXT, handle_invalid_input))

    # Add handler for file uploads
    app.add_handler(MessageHandler(filters.Document.ALL, handle_file))

    # Run the bot
    try:
        app.run_polling(timeout=30)  # Set timeout here
    except Exception as e:
        print(f"Error starting the bot: {e}")

if __name__ == "__main__":
    main()
